from __future__ import annotations

import io
import re
import tempfile
import zipfile
from pathlib import Path
from typing import Any


def _as_text(v: Any) -> str:
    if isinstance(v, dict):
        return " | ".join(f"{k}: {x}" for k, x in v.items() if x not in (None, "", [], {}))
    if isinstance(v, (list, tuple)):
        return "; ".join(_as_text(x) for x in v)
    return str(v or "")


def extract_approved_makes(corpus: str, limit: int = 150) -> list[dict[str, str]]:
    """Best-effort extraction of vendor/approved-make statements with source traceability."""
    results: list[dict[str, str]] = []
    src, page_sheet = "", ""
    trigger = re.compile(r"approved\s+(?:make|makes|manufacturer|manufacturers|vendor|vendors)|acceptable\s+(?:make|manufacturer|vendor)|make\s*[:\-]|manufacturer\s*[:\-]|vendor\s*[:\-]", re.I)
    for raw in (corpus or "").splitlines():
        line = raw.strip()
        m = re.match(r"\[SOURCE FILE:\s*([^\]|]+)", line)
        if m:
            src = m.group(1).strip(); page_sheet = ""; continue
        m = re.match(r"\[(PAGE|SHEET)[: ]+(.+?)\]", line, re.I)
        if m:
            page_sheet = f"{m.group(1).title()} {m.group(2).strip()}"; continue
        if trigger.search(line) and len(line) > 8:
            results.append({"Requirement / Statement": line[:700], "Source File": src, "Page / Sheet": page_sheet})
        if len(results) >= limit:
            break
    return results


def build_rfi_rows(analysis: dict) -> list[dict[str, Any]]:
    out = []
    items = list(analysis.get("clarifications") or []) + list(analysis.get("missing_information") or [])
    seen = set()
    for i, item in enumerate(items, 1):
        text = _as_text(item).strip()
        if not text or text.lower() in seen:
            continue
        seen.add(text.lower())
        out.append({
            "RFI No.": f"RFI-{len(out)+1:03d}",
            "Category": "Clarification" if i <= len(analysis.get("clarifications") or []) else "Missing Information",
            "Question / Information Required": text,
            "Priority": "High" if any(k in text.lower() for k in ["scope", "quantity", "drawing", "approved", "standard", "schedule"]) else "Medium",
            "Client Response": "",
            "Status": "Open",
            "Owner": "Presales",
        })
    return out


def build_responsibility_matrix(analysis: dict) -> list[dict[str, str]]:
    rows = []
    for i, scope in enumerate(analysis.get("scope_of_work") or [], 1):
        text = _as_text(scope)
        low = text.lower()
        if any(k in low for k in ["supply", "provide", "furnish"]):
            ts, client = "R/A", "C"
        elif any(k in low for k in ["install", "testing", "commission"]):
            ts, client = "R/A", "C/I"
        else:
            ts, client = "R", "C"
        rows.append({"No.": str(i), "Activity / Scope": text, "TS": ts, "Client / Main Contractor": client, "Consultant": "C/I", "Notes": "To be confirmed during tender clarification."})
    return rows


def build_scope_exclusions(analysis: dict) -> list[dict[str, str]]:
    defaults = [
        "Civil works, builder works, coring and making-good unless explicitly included in BOQ.",
        "Permanent electrical power supplies beyond dedicated system requirements unless explicitly included.",
        "Authority fees, permits and third-party inspection fees unless specifically stated.",
        "Works or quantities not shown in issued tender documents and latest revisions.",
    ]
    rows = [{"No.": str(i+1), "Proposed Exclusion / Qualification": x, "Status": "Draft - Verify"} for i, x in enumerate(defaults)]
    for x in analysis.get("missing_information") or []:
        text = _as_text(x)
        if text:
            rows.append({"No.": str(len(rows)+1), "Proposed Exclusion / Qualification": f"Pending clarification: {text}", "Status": "Open"})
    return rows


def build_requirement_solution_map(analysis: dict, compliance: list[dict]) -> list[dict[str, Any]]:
    evidence = analysis.get("evidence") or []
    rows = []
    for item in compliance:
        req = str(item.get("Requirement", ""))
        ev = ""
        src = ""
        for e in evidence:
            claim = _as_text(e.get("claim", "")) if isinstance(e, dict) else _as_text(e)
            if claim and (claim.lower() in req.lower() or req.lower()[:50] in claim.lower()):
                ev = _as_text(e.get("excerpt", "")); src = _as_text(e.get("source_file", "")); break
        rows.append({
            "No.": item.get("No.", len(rows)+1),
            "Tender Requirement": req,
            "TS Status": item.get("TS Status", ""),
            "TS Product / Solution": item.get("Matched TS Solution", ""),
            "Confidence %": item.get("Confidence %", ""),
            "Required Compliance Evidence": item.get("Action / Evidence Needed", ""),
            "Tender Evidence": ev,
            "Source": src,
            "Final Engineer Review": "Pending",
        })
    return rows


def build_tender_register_row(analysis: dict, scorecard: dict) -> list[dict[str, Any]]:
    bid = analysis.get("bid_decision") or {}
    return [{
        "Project": analysis.get("project", ""),
        "Client": analysis.get("client", ""),
        "Recommendation": scorecard.get("recommendation", bid.get("recommendation", "")),
        "Tender Score": scorecard.get("overall_score", bid.get("score", "")),
        "Technical Fit %": scorecard.get("technical_fit", ""),
        "Information Quality %": scorecard.get("information_quality", ""),
        "Risk Score %": scorecard.get("risk_score", ""),
        "Commercial Fit %": scorecard.get("commercial_fit", ""),
        "AI Confidence %": analysis.get("confidence", ""),
        "Open RFIs": len(build_rfi_rows(analysis)),
        "Risks": len(analysis.get("risks") or []),
        "Next Action": _as_text((analysis.get("next_actions") or [""])[0]),
    }]


def _write_sheet(wb, name: str, rows: list[dict[str, Any]], title: str):
    sh = wb.worksheets.add(name[:31])
    sh.get_range("A1:H1").merge()
    sh.get_range("A1").values = [[title]]
    sh.get_range("A1:H1").format = {"fill": "#16324F", "font": {"bold": True, "color": "#FFFFFF", "size": 14}, "row_height": 26}
    if not rows:
        sh.get_range("A3").values = [["No data available / requires review."]]
        return sh
    headers = list(rows[0].keys())
    matrix = [headers] + [[_as_text(r.get(h, "")) for h in headers] for r in rows]
    end_col = chr(64 + min(len(headers), 26))
    sh.get_range_by_indexes(2, 0, len(matrix), len(headers)).values = matrix
    hdr = sh.get_range_by_indexes(2, 0, 1, len(headers))
    hdr.format = {"fill": "#E9EEF3", "font": {"bold": True, "color": "#16324F"}, "wrap_text": True}
    body = sh.get_range_by_indexes(3, 0, max(1, len(rows)), len(headers))
    body.format.wrap_text = True
    sh.freeze_panes.freeze_rows(3)
    # Conservative widths: text-heavy columns remain readable without extreme autofit.
    for c in range(len(headers)):
        sh.get_range_by_indexes(2, c, len(matrix), 1).format.column_width = 18 if c < 2 else 24
    if len(headers) >= 3:
        sh.get_range_by_indexes(2, 1, len(matrix), 1).format.column_width = 36
    return sh


def build_tender_excel(analysis: dict, compliance: list[dict], scorecard: dict, corpus: str, output_path: str | Path) -> str:
    """Create the Tender Pack XLSX with openpyxl for reliable local/deployed export."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    # Remove default sheet after creating our executive summary.
    ws = wb.active
    ws.title = "Executive Summary"

    navy = "16324F"
    light = "E9EEF3"
    white = "FFFFFF"

    ws.merge_cells("A1:F1")
    ws["A1"] = "TS AI Tender & Sales Engineer – Tender Pack"
    ws["A1"].fill = PatternFill("solid", fgColor=navy)
    ws["A1"].font = Font(bold=True, color=white, size=16)
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 30

    summary = [
        ["Project", analysis.get("project", "")], ["Client", analysis.get("client", "")],
        ["Recommendation", scorecard.get("recommendation", "")], ["Tender Score", scorecard.get("overall_score", "")],
        ["AI Confidence", analysis.get("confidence", "")], ["Executive Summary", analysis.get("executive_summary", "")],
    ]
    for r, (label, value) in enumerate(summary, 3):
        ws.cell(r,1,label); ws.cell(r,2,_as_text(value))
        ws.cell(r,1).fill = PatternFill("solid", fgColor=light)
        ws.cell(r,1).font = Font(bold=True, color=navy)
        ws.cell(r,2).alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width=22; ws.column_dimensions["B"].width=65

    sheets = [
        ("Tender Register", build_tender_register_row(analysis, scorecard), "TS Tender Register"),
        ("Compliance", compliance, "Detailed Compliance Matrix"),
        ("Requirement Map", build_requirement_solution_map(analysis, compliance), "Requirement ↔ TS Solution ↔ Evidence"),
        ("RFI", build_rfi_rows(analysis), "Clarification / RFI Register"),
        ("Responsibility", build_responsibility_matrix(analysis), "Responsibility Matrix (RACI Draft)"),
        ("Exclusions", build_scope_exclusions(analysis), "Scope Exclusions & Qualifications"),
        ("Approved Makes", extract_approved_makes(corpus), "Approved Makes / Vendors Extraction"),
        ("BOQ Review", analysis.get("boq_summary") or [], "BOQ Review"),
        ("Risks", [{"No.": i+1, "Risk": _as_text(x), "Mitigation / Action": "Review and assign owner"} for i, x in enumerate(analysis.get("risks") or [])], "Risk Register"),
    ]

    for name, rows, title in sheets:
        sh = wb.create_sheet(title=name[:31])
        sh.merge_cells("A1:H1")
        sh["A1"] = title
        sh["A1"].fill = PatternFill("solid", fgColor=navy)
        sh["A1"].font = Font(bold=True, color=white, size=14)
        sh.row_dimensions[1].height = 26
        if not rows:
            sh["A3"] = "No data available / requires review."
            continue
        headers = list(rows[0].keys())
        for c,h in enumerate(headers,1):
            cell=sh.cell(3,c,_as_text(h))
            cell.fill=PatternFill("solid", fgColor=light)
            cell.font=Font(bold=True,color=navy)
            cell.alignment=Alignment(wrap_text=True,vertical="top")
        for r_idx,row in enumerate(rows,4):
            for c_idx,h in enumerate(headers,1):
                sh.cell(r_idx,c_idx,_as_text(row.get(h,""))).alignment=Alignment(wrap_text=True,vertical="top")
        sh.freeze_panes="A4"
        sh.auto_filter.ref=f"A3:{get_column_letter(len(headers))}{3+len(rows)}"
        for c in range(1,len(headers)+1):
            sh.column_dimensions[get_column_letter(c)].width = 36 if c==2 else (18 if c<=2 else 24)

    out = str(output_path)
    wb.save(out)
    return out

def build_tender_csv_zip(analysis: dict, compliance: list[dict], scorecard: dict, corpus: str, output_path: str | Path) -> str:
    """Portable fallback export if artifact_tool is unavailable in an external deployment."""
    import csv
    datasets = {
        "Tender_Register.csv": build_tender_register_row(analysis, scorecard),
        "Compliance.csv": compliance,
        "Requirement_Map.csv": build_requirement_solution_map(analysis, compliance),
        "RFI.csv": build_rfi_rows(analysis),
        "Responsibility.csv": build_responsibility_matrix(analysis),
        "Exclusions.csv": build_scope_exclusions(analysis),
        "Approved_Makes.csv": extract_approved_makes(corpus),
        "BOQ_Review.csv": analysis.get("boq_summary") or [],
    }
    with zipfile.ZipFile(output_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for fname, rows in datasets.items():
            s = io.StringIO()
            if rows:
                w = csv.DictWriter(s, fieldnames=list(rows[0].keys())); w.writeheader(); w.writerows(rows)
            zf.writestr(fname, '\ufeff' + s.getvalue())
    return str(output_path)
