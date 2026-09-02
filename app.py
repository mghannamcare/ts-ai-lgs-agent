import os, tempfile, math, json, sqlite3, uuid
from pathlib import Path
from datetime import datetime
import pandas as pd
import streamlit as st
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak

from ts_engine.document_parser import parse_document, image_to_data_url, pdf_pages_as_data_urls
from ts_engine.analyzer import extract_image_text_with_ai
from ts_engine.lgs_estimator import (
    price_database, parametric_takeoff, summarize, ai_lgs_takeoff, price_ai_rows,
    benchmark_project_1_rows, benchmark_project_1_meta
)

load_dotenv()
st.set_page_config(page_title="TS AI – LGS Quantity & Cost Estimator", page_icon="🏗️", layout="wide")

st.markdown("""<style>
:root{--ts-navy:#0B1F33;--ts-teal:#0F766E;--ts-teal2:#14B8A6;--ts-bg:#F4F7FA;--ts-card:#FFFFFF;--ts-text:#172033;--ts-muted:#667085;--ts-border:#D8E1EA}
.stApp{background:var(--ts-bg);color:var(--ts-text)}
.block-container{padding-top:1.25rem;max-width:1500px;padding-bottom:2.5rem}
.hero{padding:26px 30px;border:0;border-radius:18px;margin-bottom:16px;background:linear-gradient(135deg,#0B1F33 0%,#123B50 58%,#0F766E 100%);box-shadow:0 10px 30px rgba(11,31,51,.16)}
.hero h1{margin:0;color:white;font-size:2.25rem;font-weight:800;letter-spacing:-.02em}
.hero p{margin:8px 0 0;color:#D7F3EE;font-size:1.02rem}
[data-testid="stSidebar"]{background:#0B1F33;border-right:1px solid #16354F}
[data-testid="stSidebar"] *{color:#F8FAFC}
[data-testid="stSidebar"] [data-testid="stMarkdownContainer"] p{color:#D6E0EA}
[data-testid="stSidebar"] hr{border-color:#29435B}
[data-testid="stSidebar"] button{background:#0F766E!important;color:white!important;border:1px solid #21A69A!important;border-radius:10px!important}
[data-testid="stSidebar"] button:hover{background:#0B625C!important}
[data-testid="stMetric"]{background:var(--ts-card);border:1px solid var(--ts-border);padding:14px 16px;border-radius:14px;box-shadow:0 3px 12px rgba(16,24,40,.06)}
[data-testid="stMetricLabel"]{color:var(--ts-muted)!important;font-weight:650}
[data-testid="stMetricValue"]{color:var(--ts-navy)!important;font-weight:800}
.stTabs [data-baseweb="tab-list"]{gap:8px;background:#E9EEF3;padding:6px;border-radius:12px}
.stTabs [data-baseweb="tab"]{height:42px;border-radius:9px;padding:0 16px;color:#344054;font-weight:650}
.stTabs [aria-selected="true"]{background:white!important;color:var(--ts-teal)!important;box-shadow:0 2px 8px rgba(16,24,40,.08)}
.stButton>button,.stDownloadButton>button{border-radius:10px;font-weight:700;border:1px solid var(--ts-teal)}
.stButton>button[kind="primary"]{background:var(--ts-teal);color:white;border-color:var(--ts-teal)}
.stButton>button[kind="primary"]:hover{background:#0B625C;border-color:#0B625C}
[data-testid="stFileUploader"]{background:white;border:1px solid var(--ts-border);border-radius:14px;padding:6px 10px}
[data-testid="stDataFrame"], [data-testid="stDataEditor"]{background:white;border-radius:12px;overflow:hidden;border:1px solid var(--ts-border)}
div[data-testid="stAlert"]{border-radius:12px;border-width:1px}
.stTextInput input,.stNumberInput input,.stTextArea textarea{background:white!important;color:var(--ts-text)!important;border-radius:9px!important}
h1,h2,h3{color:var(--ts-navy)}
.caption,.stCaption{color:var(--ts-muted)!important}
.ts-badge{display:inline-block;margin-top:12px;padding:5px 10px;border-radius:999px;background:rgba(255,255,255,.13);color:#E6FFFA;font-size:.78rem;font-weight:700;letter-spacing:.03em}
.quote-box{background:#fff;border:1px solid #D8E1EA;border-radius:14px;padding:18px 20px;margin:8px 0 16px}
</style>""", unsafe_allow_html=True)

st.markdown('<div class="hero"><h1>TS AI – LGS Quantity & Cost Estimator</h1><p>Drawings + Specifications + BOQ → Detailed LGS Quantity Takeoff → Labor Hours → Commercial Quotation</p><span class="ts-badge">AI AGENT · LGS ESTIMATING · SAUDI MARKET</span></div>', unsafe_allow_html=True)
st.warning("Cost rates are indicative Saudi-market baselines for estimating only. Confirm supplier/subcontractor quotations before issuing a commercial offer.")

# -----------------------------
# Project/version persistence
# -----------------------------
DB_PATH = Path("ts_projects.db")

def db_conn():
    con = sqlite3.connect(DB_PATH)
    con.row_factory = sqlite3.Row
    return con


def init_db():
    with db_conn() as con:
        con.execute("""
        CREATE TABLE IF NOT EXISTS lgs_projects (
            project_id TEXT PRIMARY KEY,
            project_name TEXT NOT NULL,
            client TEXT,
            location TEXT,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        )""")
        con.execute("""
        CREATE TABLE IF NOT EXISTS lgs_project_versions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id TEXT NOT NULL,
            version_no INTEGER NOT NULL,
            saved_at TEXT NOT NULL,
            project_name TEXT NOT NULL,
            client TEXT,
            location TEXT,
            area_m2 REAL DEFAULT 0,
            rows_json TEXT NOT NULL,
            ai_result_json TEXT,
            corpus TEXT,
            assumptions_json TEXT,
            commercial_json TEXT,
            UNIQUE(project_id, version_no)
        )""")
        con.commit()


def list_projects():
    with db_conn() as con:
        return con.execute("""
        SELECT p.project_id, p.project_name, p.client, p.location, p.updated_at,
               COALESCE(MAX(v.version_no),0) AS latest_version
        FROM lgs_projects p
        LEFT JOIN lgs_project_versions v ON v.project_id=p.project_id
        GROUP BY p.project_id
        ORDER BY p.updated_at DESC
        """).fetchall()


def list_versions(project_id):
    with db_conn() as con:
        return con.execute("""
        SELECT id, version_no, saved_at, project_name, client, location, area_m2
        FROM lgs_project_versions WHERE project_id=? ORDER BY version_no DESC
        """, (project_id,)).fetchall()


def save_new_version(project_id, project_name, client, location, area_m2, rows, ai_result, corpus, assumptions, commercial):
    now = datetime.now().isoformat(timespec="seconds")
    if not project_id:
        project_id = str(uuid.uuid4())
    with db_conn() as con:
        existing = con.execute("SELECT project_id FROM lgs_projects WHERE project_id=?", (project_id,)).fetchone()
        if existing:
            con.execute("UPDATE lgs_projects SET project_name=?,client=?,location=?,updated_at=? WHERE project_id=?",
                        (project_name, client, location, now, project_id))
        else:
            con.execute("INSERT INTO lgs_projects(project_id,project_name,client,location,created_at,updated_at) VALUES(?,?,?,?,?,?)",
                        (project_id, project_name, client, location, now, now))
        latest = con.execute("SELECT COALESCE(MAX(version_no),0) FROM lgs_project_versions WHERE project_id=?", (project_id,)).fetchone()[0]
        version_no = int(latest) + 1
        con.execute("""
        INSERT INTO lgs_project_versions(project_id,version_no,saved_at,project_name,client,location,area_m2,rows_json,ai_result_json,corpus,assumptions_json,commercial_json)
        VALUES(?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            project_id, version_no, now, project_name, client, location, float(area_m2 or 0),
            json.dumps(rows, ensure_ascii=False, default=str),
            json.dumps(ai_result or {}, ensure_ascii=False, default=str),
            corpus or "",
            json.dumps(assumptions or {}, ensure_ascii=False, default=str),
            json.dumps(commercial or {}, ensure_ascii=False, default=str)
        ))
        con.commit()
    return project_id, version_no


def load_version(version_id):
    with db_conn() as con:
        return con.execute("SELECT * FROM lgs_project_versions WHERE id=?", (version_id,)).fetchone()

init_db()

# -----------------------------
# Session defaults
# -----------------------------
def ss_default(key, value):
    if key not in st.session_state:
        st.session_state[key] = value

ss_default("rows", [])
ss_default("ai_result", {})
ss_default("corpus", "")
ss_default("current_project_id", None)
ss_default("project_name", "LGS Building Estimate")
ss_default("client_name", "")
ss_default("project_location", "")
ss_default("project_area_m2", 0.0)
ss_default("detected_area_m2", 0.0)
ss_default("overhead_pct", 10.0)
ss_default("risk_pct", 5.0)
ss_default("profit_pct", 15.0)
ss_default("equipment_pct", 2.0)
ss_default("mep_rate", 450.0)
ss_default("trailer_rate", 0.0)
ss_default("vat_pct", 15.0)


def reset_new_project():
    st.session_state.current_project_id = None
    st.session_state.project_name = "New LGS Project"
    st.session_state.client_name = ""
    st.session_state.project_location = ""
    st.session_state.project_area_m2 = 0.0
    st.session_state.detected_area_m2 = 0.0
    st.session_state.rows = []
    st.session_state.ai_result = {}
    st.session_state.corpus = ""
    st.session_state.pop("xlsx_bytes", None)
    st.session_state.pop("quotation_pdf", None)


def load_demo_project():
    meta = benchmark_project_1_meta()
    st.session_state.rows = benchmark_project_1_rows()
    st.session_state.ai_result = meta
    st.session_state.corpus = "Demo Project - SRA Contractor Office loaded from supplied SRA layout + BOQ."
    st.session_state.current_project_id = None
    st.session_state.project_name = "SRA Contractor Office"
    st.session_state.client_name = "SRA"
    st.session_state.project_location = "Jeddah - KAIA"
    st.session_state.project_area_m2 = float(meta.get("area_m2") or 365.64)
    st.session_state.detected_area_m2 = float(meta.get("area_m2") or 365.64)


def load_version_into_session(version_id):
    rec = load_version(version_id)
    if not rec:
        return
    st.session_state.current_project_id = rec["project_id"]
    st.session_state.project_name = rec["project_name"]
    st.session_state.client_name = rec["client"] or ""
    st.session_state.project_location = rec["location"] or ""
    st.session_state.project_area_m2 = float(rec["area_m2"] or 0)
    st.session_state.detected_area_m2 = float(rec["area_m2"] or 0)
    st.session_state.rows = json.loads(rec["rows_json"] or "[]")
    st.session_state.ai_result = json.loads(rec["ai_result_json"] or "{}")
    st.session_state.corpus = rec["corpus"] or ""
    commercial = json.loads(rec["commercial_json"] or "{}")
    for k in ["overhead_pct","risk_pct","profit_pct","equipment_pct","mep_rate","trailer_rate","vat_pct"]:
        if k in commercial:
            st.session_state[k] = float(commercial[k])


def effective_area_m2():
    manual = float(st.session_state.get("project_area_m2", 0) or 0)
    if manual > 0:
        return manual
    ai_area = float((st.session_state.get("ai_result") or {}).get("area_m2") or 0)
    if ai_area > 0:
        return ai_area
    return float(st.session_state.get("detected_area_m2", 0) or 0)


def commercial_summary(rows, area_m2, equipment_pct, mep_rate, trailer_rate, overhead_pct, risk_pct, profit_pct, vat_pct):
    clean = [r for r in rows if pd.notna(r.get("Material Cost SAR"))]
    base = summarize(clean)
    direct = float(base["direct_cost_sar"])
    equipment = direct * float(equipment_pct) / 100.0
    mep = float(area_m2 or 0) * float(mep_rate or 0)
    module_area = 12.0 * 3.6
    trailers = int(math.ceil(float(area_m2 or 0) / module_area)) if float(area_m2 or 0) > 0 else 0
    shipping = trailers * float(trailer_rate or 0)
    cost_before_oh_risk = direct + equipment + mep + shipping
    overhead = cost_before_oh_risk * float(overhead_pct) / 100.0
    risk = cost_before_oh_risk * float(risk_pct) / 100.0
    cost_after_oh_risk = cost_before_oh_risk + overhead + risk
    profit = cost_after_oh_risk * float(profit_pct) / 100.0
    selling_excl_vat = cost_after_oh_risk + profit
    vat = selling_excl_vat * float(vat_pct) / 100.0
    selling_incl_vat = selling_excl_vat + vat
    return {
        **base,
        "equipment_consumables_sar": round(equipment, 2),
        "mep_provisional_sar": round(mep, 2),
        "module_area_m2": module_area,
        "trailer_count": trailers,
        "trailer_rate_sar": round(float(trailer_rate or 0), 2),
        "shipping_sar": round(shipping, 2),
        "cost_before_oh_risk_sar": round(cost_before_oh_risk, 2),
        "overhead_pct": float(overhead_pct), "overhead_sar": round(overhead, 2),
        "risk_pct": float(risk_pct), "risk_sar": round(risk, 2),
        "profit_pct": float(profit_pct), "profit_sar": round(profit, 2),
        "selling_price_excl_vat_sar": round(selling_excl_vat, 2),
        "vat_pct": float(vat_pct), "vat_sar": round(vat, 2),
        "selling_price_incl_vat_sar": round(selling_incl_vat, 2)
    }


def customer_quote_lines(rows, summary):
    df = pd.DataFrame(rows)
    if df.empty:
        return []
    df["Total Direct Cost SAR"] = pd.to_numeric(df.get("Total Direct Cost SAR"), errors="coerce").fillna(0)
    grouped = df.groupby("Category", dropna=False)["Total Direct Cost SAR"].sum().reset_index()
    base_cost = float(summary["cost_before_oh_risk_sar"] or 0)
    selling = float(summary["selling_price_excl_vat_sar"] or 0)
    uplift = selling / base_cost if base_cost else 1.0
    lines = []
    for _, r in grouped.iterrows():
        amount = float(r["Total Direct Cost SAR"] or 0) * uplift
        if amount > 0:
            lines.append((str(r["Category"] or "LGS Works"), amount))
    extras = [
        ("Equipment / Consumables", float(summary["equipment_consumables_sar"])),
        ("MEP Provisional Allowance", float(summary["mep_provisional_sar"])),
        (f"Transport / Delivery ({summary['trailer_count']} trailer(s))", float(summary["shipping_sar"]))
    ]
    for label, val in extras:
        if val > 0:
            lines.append((label, val * uplift))
    # Normalize minor floating difference to exact selling subtotal.
    if lines:
        diff = selling - sum(v for _, v in lines)
        label, val = lines[-1]
        lines[-1] = (label, val + diff)
    return lines


def make_quotation_pdf(project_name, client, location, area_m2, rows, summary, version_label="Current"):
    out = Path(tempfile.gettempdir()) / "TS_AI_Commercial_Quotation.pdf"
    doc = SimpleDocTemplate(str(out), pagesize=A4, rightMargin=16*mm, leftMargin=16*mm, topMargin=14*mm, bottomMargin=14*mm)
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="TSHeader", parent=styles["Title"], fontName="Helvetica-Bold", fontSize=20, leading=24, textColor=colors.HexColor("#0B1F33"), alignment=TA_LEFT, spaceAfter=6))
    styles.add(ParagraphStyle(name="TSTag", parent=styles["Normal"], fontName="Helvetica-Bold", fontSize=9, textColor=colors.HexColor("#0F766E"), spaceAfter=12))
    styles.add(ParagraphStyle(name="TSNote", parent=styles["Normal"], fontSize=8.5, leading=12, textColor=colors.HexColor("#475467")))
    story = [
        Paragraph("TS AI - Commercial Quotation", styles["TSHeader"]),
        Paragraph("LGS / Prefabricated Building Estimate", styles["TSTag"]),
    ]
    info = [
        ["Project", project_name or "-", "Version", version_label],
        ["Client", client or "-", "Date", datetime.now().strftime("%d-%b-%Y")],
        ["Location", location or "-", "Area", f"{float(area_m2 or 0):,.2f} m2"],
    ]
    t = Table(info, colWidths=[25*mm, 73*mm, 22*mm, 48*mm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,-1), colors.HexColor("#F6F8FA")),
        ("TEXTCOLOR", (0,0), (-1,-1), colors.HexColor("#172033")),
        ("FONTNAME", (0,0), (-1,-1), "Helvetica"),
        ("FONTNAME", (0,0), (0,-1), "Helvetica-Bold"),
        ("FONTNAME", (2,0), (2,-1), "Helvetica-Bold"),
        ("FONTSIZE", (0,0), (-1,-1), 9),
        ("GRID", (0,0), (-1,-1), 0.4, colors.HexColor("#D8E1EA")),
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("BOTTOMPADDING", (0,0), (-1,-1), 7), ("TOPPADDING", (0,0), (-1,-1), 7),
    ]))
    story += [t, Spacer(1, 8*mm), Paragraph("Commercial Price Breakdown", styles["Heading2"])]
    quote_lines = customer_quote_lines(rows, summary)
    data = [["No.", "Description", "Amount (SAR)"]]
    for i, (label, amount) in enumerate(quote_lines, 1):
        data.append([str(i), label, f"{amount:,.2f}"])
    data.append(["", "Subtotal excl. VAT", f"{summary['selling_price_excl_vat_sar']:,.2f}"])
    data.append(["", f"VAT {summary['vat_pct']:.1f}%", f"{summary['vat_sar']:,.2f}"])
    data.append(["", "TOTAL incl. VAT", f"{summary['selling_price_incl_vat_sar']:,.2f}"])
    qt = Table(data, colWidths=[13*mm, 117*mm, 38*mm], repeatRows=1)
    qt.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#0B1F33")),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTNAME", (0,-3), (-1,-1), "Helvetica-Bold"),
        ("BACKGROUND", (0,-1), (-1,-1), colors.HexColor("#DFF7F2")),
        ("ALIGN", (0,0), (0,-1), "CENTER"),
        ("ALIGN", (2,1), (2,-1), "RIGHT"),
        ("GRID", (0,0), (-1,-1), 0.45, colors.HexColor("#D0D5DD")),
        ("FONTSIZE", (0,0), (-1,-1), 9),
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("TOPPADDING", (0,0), (-1,-1), 7), ("BOTTOMPADDING", (0,0), (-1,-1), 7),
    ]))
    story += [qt, Spacer(1, 7*mm)]
    logistics = f"Logistics basis: module footprint 12.0 m x 3.6 m = {summary['module_area_m2']:.2f} m2/module. Estimated transport requirement: {summary['trailer_count']} trailer(s)."
    story += [Paragraph(logistics, styles["TSNote"]), Spacer(1, 3*mm)]
    story += [Paragraph("Commercial Notes", styles["Heading3"]),
              Paragraph("1. This quotation is generated from the current TS AI estimate and is subject to final drawings, approved specifications, structural design, site conditions and supplier/subcontractor quotations.", styles["TSNote"]),
              Paragraph("2. Quantities and rates remain editable in the estimator. Any commercial revision should be saved as a new project version before issuing a revised quotation.", styles["TSNote"]),
              Paragraph("3. Internal overhead, risk and profit percentages are not shown in this client-facing quotation.", styles["TSNote"])]
    doc.build(story)
    return out

# -----------------------------
# Sidebar
# -----------------------------
with st.sidebar:
    st.subheader("Engine")
    st.write("Document parser: Ready")
    st.write("AI/Vision:", "Ready" if os.getenv("OPENAI_API_KEY") else "API key required for drawing intelligence")
    st.caption("PDF, DOCX, XLS/XLSX, CSV and images are supported. CAD/BIM direct reading will be added later; export DWG/Revit to PDF for the current workflow.")
    st.divider()

    st.subheader("Project Information")
    c1, c2 = st.columns(2)
    c1.button("New Project", use_container_width=True, on_click=reset_new_project)
    c2.button("Load Demo", use_container_width=True, on_click=load_demo_project)
    st.text_input("Project Name", key="project_name")
    st.text_input("Client", key="client_name")
    st.text_input("Project Location", key="project_location")
    st.number_input("Building Area (m²)", min_value=0.0, max_value=1000000.0, step=1.0, key="project_area_m2", help="Enter actual building footprint area. If left at 0, detected/parametric area is used where available.")

    st.divider()
    st.subheader("Commercial Settings")
    st.number_input("Overhead %", 0.0, 50.0, step=0.5, key="overhead_pct")
    st.number_input("Risk %", 0.0, 50.0, step=0.5, key="risk_pct")
    st.number_input("Profit %", 0.0, 100.0, step=0.5, key="profit_pct", help="Applied as markup on cost after overhead and risk.")
    st.number_input("Equipment / consumables %", 0.0, 30.0, step=0.5, key="equipment_pct")
    st.number_input("MEP provisional SAR/m²", 0.0, 2500.0, step=25.0, key="mep_rate")
    st.number_input("Trailer shipping cost (SAR / trailer)", 0.0, 100000.0, step=500.0, key="trailer_rate")
    st.number_input("VAT %", 0.0, 20.0, step=0.5, key="vat_pct")
    module_area = 12.0 * 3.6
    sidebar_area = effective_area_m2()
    tcount = int(math.ceil(sidebar_area / module_area)) if sidebar_area > 0 else 0
    st.caption(f"Transport basis: 12.0 × 3.6 m = {module_area:.1f} m²/module → {tcount} trailer(s) estimated.")

    st.divider()
    st.subheader("Projects & Versions")
    projects = list_projects()
    if projects:
        labels = {f"{r['project_name']} | {r['client'] or '-'} | v{r['latest_version']}": r['project_id'] for r in projects}
        selected_label = st.selectbox("Saved Projects", ["-- Select --"] + list(labels.keys()))
        if selected_label != "-- Select --":
            pid = labels[selected_label]
            versions = list_versions(pid)
            vlabels = {f"v{r['version_no']} - {r['saved_at'].replace('T',' ')}": r['id'] for r in versions}
            selected_vlabel = st.selectbox("Version", list(vlabels.keys()))
            st.button("Load Selected Version", use_container_width=True, on_click=load_version_into_session, args=(vlabels[selected_vlabel],))
        st.caption("Each save creates a new version; previous versions are never overwritten.")
    else:
        st.caption("No saved LGS estimates yet.")
    st.caption("Cloud note: Streamlit Community Cloud local SQLite storage may reset after redeploy/restart. For permanent production history, connect a cloud database later.")

# -----------------------------
# Main tabs
# -----------------------------
tab1, tab2, tab3, tab4, tab5 = st.tabs([
    "1. Upload & Extract", "2. Quantity Takeoff", "3. Cost & Selling Price", "4. Commercial Quotation", "5. Price Database"
])

with tab1:
    st.subheader("Project Documents")
    st.write(f"**Project:** {st.session_state.project_name}  |  **Client:** {st.session_state.client_name or '-'}  |  **Location:** {st.session_state.project_location or '-'}")
    files = st.file_uploader("Upload drawings, specifications and BOQ", accept_multiple_files=True,
        type=["pdf","docx","xlsx","xls","csv","png","jpg","jpeg","webp"])
    if st.button("Read Project Files", type="primary", use_container_width=True):
        corpus=[]; status=[]
        for f in files or []:
            suffix=Path(f.name).suffix
            tmp=Path(tempfile.gettempdir())/(f"ts_lgs_{abs(hash(f.name))}{suffix}")
            tmp.write_bytes(f.getvalue())
            try:
                parsed=parse_document(tmp); text=parsed.get("text","")
                if parsed.get("needs_vision") and os.getenv("OPENAI_API_KEY"):
                    urls=pdf_pages_as_data_urls(tmp,8) if suffix.lower()==".pdf" else [image_to_data_url(tmp)]
                    text += "\n"+extract_image_text_with_ai(urls,"Focus on dimensions, wall/roof/floor build-ups, openings, LGS framing notes, material specifications and quantities.")
                corpus.append(f"\n[SOURCE FILE: {f.name}]\n{text}")
                status.append({"File":f.name,"Type":parsed.get("file_type"),"Status":"Read","Needs Vision":parsed.get("needs_vision",False)})
            except Exception as e:
                status.append({"File":f.name,"Status":"Error","Details":str(e)})
        st.session_state.corpus="\n".join(corpus)
        st.dataframe(pd.DataFrame(status),use_container_width=True,hide_index=True)
        st.success(f"Files processed. Extracted corpus: {len(st.session_state.corpus):,} characters.")
    if st.session_state.corpus:
        with st.expander("Extracted source text"):
            st.text_area("Source",st.session_state.corpus[:120000],height=350)

with tab2:
    st.subheader("A. AI Takeoff from Uploaded Documents")
    if st.button("Generate LGS Takeoff from Documents",use_container_width=True):
        if not st.session_state.corpus:
            st.error("Upload and read project files first.")
        elif not os.getenv("OPENAI_API_KEY"):
            st.error("OPENAI_API_KEY is required for intelligent drawing/specification takeoff.")
        else:
            with st.spinner("Analyzing LGS scope..."):
                res=ai_lgs_takeoff(st.session_state.corpus,st.session_state.project_name)
                st.session_state.ai_result=res or {}
                if (res or {}).get("area_m2"):
                    st.session_state.detected_area_m2 = float((res or {}).get("area_m2") or 0)
                st.session_state.rows=price_ai_rows((res or {}).get("quantity_rows",[]))
            st.success("LGS takeoff generated.")
    if st.session_state.ai_result:
        r=st.session_state.ai_result
        if r.get("dimensions"): st.write("**Benchmark dimensions:**", r["dimensions"], " | Area:", r.get("area_m2"), "m²")
        if r.get("building_dimensions"): st.write("**Detected dimensions:**",r["building_dimensions"])
        c1,c2=st.columns(2)
        with c1:
            st.write("**Assumptions**")
            for x in r.get("assumptions",[]): st.write("•",x)
        with c2:
            st.write("**Clarifications required**")
            for x in r.get("clarifications",[]): st.write("•",x)

    st.divider()
    st.subheader("B. Parametric Fallback / Quick Estimate")
    c1,c2,c3,c4=st.columns(4)
    L=c1.number_input("Length (m)",1.0,200.0,12.0,.5); W=c2.number_input("Width (m)",1.0,100.0,6.0,.5)
    H=c3.number_input("Wall height (m)",2.0,12.0,3.0,.1); P=c4.number_input("Internal partition length (m)",0.0,500.0,18.0,1.0)
    c1,c2,c3,c4=st.columns(4)
    ed=c1.number_input("External doors",0,100,2); idr=c2.number_input("Internal doors",0,200,4)
    wa=c3.number_input("Window area (m²)",0.0,500.0,12.0,1.0); steel=c4.number_input("LGS steel intensity (kg/m²)",15.0,90.0,38.0,1.0)
    if st.button("Generate Parametric Takeoff",use_container_width=True):
        st.session_state.rows=parametric_takeoff(L,W,H,P,ed,idr,wa,steel)
        st.session_state.detected_area_m2 = round(L*W,2)
        st.session_state.ai_result={"area_m2":round(L*W,2),"assumptions":["Parametric estimate generated from entered dimensions. Verify against IFC/shop drawings."],"clarifications":[]}
        st.success("Parametric takeoff generated and project area updated.")

    if st.session_state.rows:
        st.subheader("Detailed Quantity Takeoff")
        df=pd.DataFrame(st.session_state.rows)
        edited=st.data_editor(df,use_container_width=True,hide_index=True,num_rows="dynamic",key="takeoff_editor")
        st.session_state.rows=edited.to_dict("records")

with tab3:
    rows=st.session_state.rows
    if not rows:
        st.info("Generate a takeoff first.")
    else:
        total = commercial_summary(
            rows, effective_area_m2(), st.session_state.equipment_pct, st.session_state.mep_rate,
            st.session_state.trailer_rate, st.session_state.overhead_pct, st.session_state.risk_pct,
            st.session_state.profit_pct, st.session_state.vat_pct
        )
        c=st.columns(5)
        c[0].metric("Material Cost",f"{total['material_cost_sar']:,.0f} SAR")
        c[1].metric("Labor Hours",f"{total['labor_hours']:,.0f} h")
        c[2].metric("Direct Cost",f"{total['direct_cost_sar']:,.0f} SAR")
        c[3].metric("Shipping",f"{total['shipping_sar']:,.0f} SAR",f"{total['trailer_count']} trailers")
        c[4].metric("Selling Price excl. VAT",f"{total['selling_price_excl_vat_sar']:,.0f} SAR")

        st.subheader("Commercial Cost Build-up")
        summary_df = pd.DataFrame([
            ["Materials", total["material_cost_sar"]], ["Labor", total["labor_cost_sar"]],
            ["Equipment / Consumables", total["equipment_consumables_sar"]], ["MEP Provisional", total["mep_provisional_sar"]],
            [f"Shipping ({total['trailer_count']} trailers × {total['trailer_rate_sar']:,.0f})", total["shipping_sar"]],
            [f"Overhead ({total['overhead_pct']:.1f}%)", total["overhead_sar"]], [f"Risk ({total['risk_pct']:.1f}%)", total["risk_sar"]],
            [f"Profit ({total['profit_pct']:.1f}%)", total["profit_sar"]], ["Selling Price excl. VAT", total["selling_price_excl_vat_sar"]],
            [f"VAT ({total['vat_pct']:.1f}%)", total["vat_sar"]], ["Grand Total incl. VAT", total["selling_price_incl_vat_sar"]]
        ], columns=["Component","SAR"])
        st.dataframe(summary_df, use_container_width=True, hide_index=True)
        st.caption(f"Trailer calculation: ceil({effective_area_m2():,.2f} m² ÷ 43.20 m² per 12×3.6 m module) = {total['trailer_count']} trailers.")

        unpriced=[r for r in rows if pd.isna(r.get("Material Cost SAR"))]
        if unpriced: st.warning(f"{len(unpriced)} item(s) require supplier pricing before the estimate is complete.")

        commercial_settings = {
            "overhead_pct": st.session_state.overhead_pct, "risk_pct": st.session_state.risk_pct,
            "profit_pct": st.session_state.profit_pct, "equipment_pct": st.session_state.equipment_pct,
            "mep_rate": st.session_state.mep_rate, "trailer_rate": st.session_state.trailer_rate, "vat_pct": st.session_state.vat_pct
        }
        if st.button("Save as New Project Version", type="primary", use_container_width=True):
            if not st.session_state.project_name.strip():
                st.error("Enter a Project Name first.")
            else:
                pid, vno = save_new_version(
                    st.session_state.current_project_id, st.session_state.project_name, st.session_state.client_name,
                    st.session_state.project_location, effective_area_m2(), rows, st.session_state.ai_result,
                    st.session_state.corpus, {}, commercial_settings
                )
                st.session_state.current_project_id = pid
                st.success(f"Saved as version v{vno}. Previous versions remain unchanged.")

        def make_xlsx():
            wb=Workbook(); ws=wb.active; ws.title="LGS Cost Estimate"
            headers=list(pd.DataFrame(rows).columns); ws.append(headers)
            for cell in ws[1]:
                cell.font=Font(bold=True,color="FFFFFF"); cell.fill=PatternFill("solid",fgColor="17365D"); cell.alignment=Alignment(horizontal="center")
            for r in rows: ws.append([r.get(h) for h in headers])
            ws.freeze_panes="A2"; ws.auto_filter.ref=ws.dimensions
            for col in ws.columns:
                letter=col[0].column_letter; ws.column_dimensions[letter].width=min(38,max(12,max(len(str(c.value or "")) for c in col)+2))
            s=wb.create_sheet("Commercial Summary"); s.append(["Project",st.session_state.project_name]); s.append(["Client",st.session_state.client_name]); s.append(["Location",st.session_state.project_location]); s.append(["Area m2",effective_area_m2()]); s.append([]); s.append(["Cost Component","SAR"])
            for _, rr in summary_df.iterrows(): s.append([rr["Component"],float(rr["SAR"])])
            p=wb.create_sheet("Price Database"); pdb=price_database(); ph=list(pdb[0].keys()); p.append(ph)
            for x in pdb:p.append([x.get(h) for h in ph])
            out=Path(tempfile.gettempdir())/"TS_LGS_Detailed_Cost_Estimate.xlsx"; wb.save(out); return out
        if st.button("Prepare Excel Cost Estimate", use_container_width=True):
            out=make_xlsx(); st.session_state["xlsx_bytes"]=out.read_bytes()
        if st.session_state.get("xlsx_bytes"):
            st.download_button("Download TS LGS Cost Estimate Excel",st.session_state["xlsx_bytes"],"TS_LGS_Detailed_Cost_Estimate.xlsx",
                               "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",use_container_width=True)

with tab4:
    rows = st.session_state.rows
    if not rows:
        st.info("Generate a takeoff and cost estimate first.")
    else:
        total = commercial_summary(
            rows, effective_area_m2(), st.session_state.equipment_pct, st.session_state.mep_rate,
            st.session_state.trailer_rate, st.session_state.overhead_pct, st.session_state.risk_pct,
            st.session_state.profit_pct, st.session_state.vat_pct
        )
        st.subheader("Client-Facing Commercial Quotation")
        st.markdown(f"""<div class="quote-box"><b>Project:</b> {st.session_state.project_name}<br><b>Client:</b> {st.session_state.client_name or '-'}<br><b>Location:</b> {st.session_state.project_location or '-'}<br><b>Area:</b> {effective_area_m2():,.2f} m²<br><b>Estimated trailers:</b> {total['trailer_count']} × {total['trailer_rate_sar']:,.0f} SAR</div>""", unsafe_allow_html=True)
        c1,c2,c3 = st.columns(3)
        c1.metric("Quotation excl. VAT", f"{total['selling_price_excl_vat_sar']:,.0f} SAR")
        c2.metric("VAT", f"{total['vat_sar']:,.0f} SAR")
        c3.metric("Quotation incl. VAT", f"{total['selling_price_incl_vat_sar']:,.0f} SAR")
        qdf = pd.DataFrame(customer_quote_lines(rows,total), columns=["Commercial Scope","Amount SAR"])
        st.dataframe(qdf, use_container_width=True, hide_index=True)
        st.caption("The PDF quotation does not expose internal overhead, risk or profit percentages.")
        if st.button("Generate Commercial Quotation PDF", type="primary", use_container_width=True):
            versions = list_versions(st.session_state.current_project_id) if st.session_state.current_project_id else []
            version_label = f"v{versions[0]['version_no']}" if versions else "Unsaved Draft"
            pdf = make_quotation_pdf(st.session_state.project_name, st.session_state.client_name, st.session_state.project_location,
                                     effective_area_m2(), rows, total, version_label)
            st.session_state["quotation_pdf"] = pdf.read_bytes()
        if st.session_state.get("quotation_pdf"):
            safe_name = "".join(ch if ch.isalnum() or ch in "-_" else "_" for ch in st.session_state.project_name)[:60]
            st.download_button("Download Commercial Quotation PDF", st.session_state["quotation_pdf"],
                               f"TS_AI_Quotation_{safe_name}.pdf", "application/pdf", use_container_width=True)

with tab5:
    st.subheader("Indicative Saudi Cost Baseline")
    st.caption("These are estimating baselines, not supplier quotations. Update them when current TS procurement prices are available.")
    st.dataframe(pd.DataFrame(price_database()),use_container_width=True,hide_index=True)
